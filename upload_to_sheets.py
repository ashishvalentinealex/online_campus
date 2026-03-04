from welcome_email import send_welcome_emails
from vcard_converter import create_phone_xlsx_and_vcf
import yagmail
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
import sqlite3

# Configuration
SERVICE_ACCOUNT_FILE = 'credentials.json'
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
INPUT_FILE = "newcomers_final.xlsx"
DEST_SHEET = "EFAMILY MAIN_20-10-25"
DB_FILE = 'sync_tracker.db'

def init_db():
    """Initialize database for tracking last email"""
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS last_sync (
            id INTEGER PRIMARY KEY,
            last_email TEXT,
            sync_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    return conn

def save_last_email(conn, email):
    cursor = conn.cursor()
    cursor.execute('DELETE FROM last_sync')
    cursor.execute('INSERT INTO last_sync (last_email) VALUES (?)', (email,))
    conn.commit()

def upload_to_sheets():
    # Initialize database
    conn = init_db()
    
    # Load Excel file
    print(f"Loading {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Read all data except header
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    
    print(f"Found {len(data)} records to upload")
    
    if len(data) == 0:
        print("No data to upload!")
        return
    
    # Connect to Google Sheets
    print("\nConnecting to Google Sheets...")
    credentials = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    gc = gspread.authorize(credentials)
    
    # Open destination sheet
    print(f"Opening {DEST_SHEET}...")
    dest_workbook = gc.open(DEST_SHEET)
    dest_sheet2 = dest_workbook.get_worksheet(1)  # Sheet2
    
    # Append all records
    print(f"\nUploading {len(data)} records to Sheet2...")
    dest_sheet2.append_rows(data)
    
    # Get the last email
    last_email = data[-1][0]  # Email is in first column
    
    # Save to database
    save_last_email(conn, last_email)
    
    print(f"\n✅ Successfully uploaded {len(data)} records to {DEST_SHEET} Sheet2")
    print(f"📧 Last email stored: {last_email}")
    
    conn.close()

    # Generate phone list and VCF
    print("\n📇 Generating VCF contact file...")
    create_phone_xlsx_and_vcf()

    # Send VCF to efamcare
    print("\n📨 Sending VCF to efamcare@gmail.com...")
    yag = yagmail.SMTP(user="efamcare@gmail.com", password="evkzzjrlkbepxqsm")
    yag.send(
        to="efamcare@gmail.com",
        subject="New Newcomers Contact List",
        contents=["Please find the new newcomers VCF attached.", "newcomers.vcf"]
    )
    print("✅ VCF sent!")

    # Send newcomers_final.xlsx to avation2k14@gmail.com
    print("\n📨 Sending newcomers list to avation2k14@gmail.com...")
    yag.send(
        to="avation2k14@gmail.com",
        subject="New Newcomers List",
        contents=["Please find the new newcomers list attached.", INPUT_FILE]
    )
    print("✅ Newcomers list sent!")

    # Send welcome emails to all newly synced members
    print("\n📨 Sending welcome emails...")
    send_welcome_emails(INPUT_FILE)
    print("✅ Welcome emails sent!")

if __name__ == "__main__":
    try:
        upload_to_sheets()
    except Exception as e:
        print(f"Error: {e}")
