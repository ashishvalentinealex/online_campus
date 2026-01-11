from openai import OpenAI
from openpyxl import load_workbook, Workbook
import json
from dotenv import load_dotenv
import os

# Load environment variables
load_dotenv()

# Configuration
OPENAI_API_KEY = os.getenv('OPENAI_API_KEY')
INPUT_FILE = "newcomers.xlsx"
OUTPUT_FILE = "newcomers_enriched.xlsx"

# Initialize OpenAI client
client = OpenAI(api_key=OPENAI_API_KEY)

def get_location_and_phone_info(city, phone):
    """Use OpenAI to get country, continent, and validate phone number"""
    
    prompt = f"""Given the following information:
City: {city}
Phone: {phone}

Please provide a JSON response with:
1. country: The country name for this city
2. continent: The continent name
3. phone_corrected: The phone number with proper country code (if missing, add it based on the country)

Format the response as valid JSON only, no additional text:
{{
    "country": "country name",
    "continent": "continent name",
    "phone_corrected": "phone with country code"
}}"""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that provides geographic and phone number information. Always respond with valid JSON only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        
        result = json.loads(response.choices[0].message.content)
        
        # Return result with token usage
        return result, response.usage
    except Exception as e:
        print(f"Error processing city={city}, phone={phone}: {e}")
        return {
            "country": "Unknown",
            "continent": "Unknown",
            "phone_corrected": phone
        }, None

def process_newcomers():
    # Load the input file
    print(f"Loading {INPUT_FILE}...")
    wb_input = load_workbook(INPUT_FILE)
    ws_input = wb_input.active
    
    # Create output workbook
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "Enriched Newcomers"
    
    # Add headers
    ws_output.append(["Email Address", "Name", "City", "Phone Number", "Country", "Continent"])
    
    # Process each row (skip header)
    total_rows = ws_input.max_row - 1
    print(f"Processing {total_rows} records...\n")
    
    # Token tracking
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_tokens = 0
    
    for idx, row in enumerate(ws_input.iter_rows(min_row=2, values_only=True), start=1):
        email = row[0]
        name = row[1]
        city = row[2]
        phone = row[3]
        
        print(f"[{idx}/{total_rows}] Processing: {name} - {city}")
        
        # Get enriched data from OpenAI
        info, usage = get_location_and_phone_info(city, phone)
        
        # Track tokens
        if usage:
            total_prompt_tokens += usage.prompt_tokens
            total_completion_tokens += usage.completion_tokens
            total_tokens += usage.total_tokens
        
        # Write to output
        ws_output.append([
            email,
            name,
            city,
            info['phone_corrected'],
            info['country'],
            info['continent']
        ])
        
        print(f"  → Country: {info['country']}, Continent: {info['continent']}, Phone: {info['phone_corrected']}\n")
    
    # Save output file
    wb_output.save(OUTPUT_FILE)
    print(f"\n✅ Saved enriched data to {OUTPUT_FILE}")
    print(f"Total records processed: {total_rows}")
    print(f"\n📊 Token Usage Summary:")
    print(f"  Prompt tokens: {total_prompt_tokens:,}")
    print(f"  Completion tokens: {total_completion_tokens:,}")
    print(f"  Total tokens: {total_tokens:,}")
    print(f"\n💰 Estimated Cost (gpt-4o-mini):")
    # gpt-4o-mini pricing: $0.150 per 1M input tokens, $0.600 per 1M output tokens
    input_cost = (total_prompt_tokens / 1_000_000) * 0.150
    output_cost = (total_completion_tokens / 1_000_000) * 0.600
    total_cost = input_cost + output_cost
    print(f"  Input cost: ${input_cost:.4f}")
    print(f"  Output cost: ${output_cost:.4f}")
    print(f"  Total cost: ${total_cost:.4f}")

if __name__ == "__main__":
    if not OPENAI_API_KEY:
        print("Error: Please set OPENAI_API_KEY environment variable")
        print("Example: export OPENAI_API_KEY='your-api-key-here'")
        exit(1)
    
    try:
        process_newcomers()
    except Exception as e:
        print(f"Error: {e}")
