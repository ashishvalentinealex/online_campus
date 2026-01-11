from openpyxl import load_workbook
import re

# Configuration
INPUT_FILE = "newcomers_enriched.xlsx"
OUTPUT_FILE = "newcomers_final.xlsx"

def clean_phone_number(phone):
    """Remove all non-digit characters from phone number"""
    if not phone:
        return phone
    # Keep only digits
    cleaned = re.sub(r'[^0-9]', '', str(phone))
    return cleaned

def process_phone_numbers():
    print(f"Loading {INPUT_FILE}...")
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Find the Phone Number column (should be column 4, index D)
    phone_col_idx = 4  # Column D (1-indexed)
    
    total_rows = ws.max_row - 1
    print(f"Processing {total_rows} phone numbers...\n")
    
    # Process each row (skip header)
    for row_idx in range(2, ws.max_row + 1):
        original_phone = ws.cell(row=row_idx, column=phone_col_idx).value
        cleaned_phone = clean_phone_number(original_phone)
        
        # Update the cell
        ws.cell(row=row_idx, column=phone_col_idx).value = cleaned_phone
        
        if original_phone != cleaned_phone:
            print(f"Row {row_idx-1}: {original_phone} → {cleaned_phone}")
    
    # Save the file
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Saved cleaned data to {OUTPUT_FILE}")
    print(f"Total records processed: {total_rows}")

if __name__ == "__main__":
    try:
        process_phone_numbers()
    except Exception as e:
        print(f"Error: {e}")
